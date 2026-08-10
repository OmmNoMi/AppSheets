#!/usr/bin/env python3
"""
ingest_questionnaire.py — OmmNoMi Standard AppVariables Generator
===================================================================
Generates AppVariables adhering 100% to your updated Google Sheet column schema:

Tab Name: `AppVariables` (renamed from `Variable`)

Columns (21 total):
  1.  ID
  2.  Table
  3.  Column
  4.  Tags
  5.  ValueControl
  6.  Title         (English Title / Question Prompt)
  7.  Description
  8.  UsedFor
  9.  Decimal
  10. EnumValue
  11. EnumList
  12. VariableList  (Option choices list)
  13. DateValue
  14. Photo
  15. URL
  16. File
  17. Title_hi      (Hindi Translation)
  18. Title_mr      (Marathi Translation)
  19. ActionIcon    (Emoji Icon)
  20. LastEditBy
  21. LastEditOn

This script parses questionnaire Excel matrices and emits rows matching this exact structure.
"""

import sys
import os
import argparse
import openpyxl
import csv
from datetime import datetime

STANDARD_HEADERS = [
    'ID', 'Table', 'Column', 'Tags', 'ValueControl', 'Title', 'Description', 'UsedFor',
    'Decimal', 'EnumValue', 'EnumList', 'VariableList', 'DateValue', 'Photo', 'URL', 'File',
    'Title_hi', 'Title_mr', 'ActionIcon', 'LastEditBy', 'LastEditOn'
]

def parse_questionnaire_excel(xlsx_path):
    """
    Parses questionnaire Excel sheets (Question & Variable tabs).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    questions, variables = [], []
    
    if 'Question' in wb.sheetnames:
        ws = wb['Question']
        headers = [str(cell).strip() for cell in next(ws.iter_rows(values_only=True)) if cell is not None]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if any(row):
                questions.append(dict(zip(headers, row)))
                
    if 'Variable' in wb.sheetnames or 'AppVariables' in wb.sheetnames:
        sname = 'AppVariables' if 'AppVariables' in wb.sheetnames else 'Variable'
        ws = wb[sname]
        headers = [str(cell).strip() for cell in next(ws.iter_rows(values_only=True)) if cell is not None]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if any(row):
                variables.append(dict(zip(headers, row)))

    return questions, variables


def transform_to_ommnomi_standard_appvariables(questions, variables):
    """
    Transforms questions and variables into your updated AppVariables column schema.
    """
    app_vars = []
    now_str = datetime.now().strftime("%m/%d/%Y %H:%M:%S")

    # 1. Base AppVariables (System & Company constants)
    base_constants = [
        {
            'ID': 'CompanyName', 'Table': 'AppVariable', 'Column': '', 'Tags': 'ID is used in Code , Changes on App Copy',
            'ValueControl': 'Enum', 'Title': 'Company Name', 'Description': 'OmmNoMi Survey Engine', 'UsedFor': 'Header title in reports',
            'EnumValue': 'OmmNoMi Automation LLP', 'LastEditBy': 'DevNoMi', 'LastEditOn': now_str
        },
        {
            'ID': 'AppCodeBaseFolder', 'Table': 'AppVariable', 'Column': '', 'Tags': 'ID is used in Code , Changes on App Copy',
            'ValueControl': 'URL', 'Title': 'App Base Folder', 'Description': 'Google Drive Root Folder', 'UsedFor': 'File attachments',
            'URL': 'https://drive.google.com/drive/folders/ommnomi', 'LastEditBy': 'DevNoMi', 'LastEditOn': now_str
        }
    ]
    for bc in base_constants:
        row = {h: '' for h in STANDARD_HEADERS}
        row.update(bc)
        app_vars.append(row)

    # 2. Standalone Option Item Variables
    for v in variables:
        var_id = str(v.get('ID', '')).strip()
        if not var_id:
            continue
        title_en = v.get('Name_en', v.get('Title', var_id))
        title_hi = v.get('Name_hi', v.get('Title_hi', ''))
        title_mr = v.get('Name_mr', v.get('Title_mr', ''))
        row = {h: '' for h in STANDARD_HEADERS}
        row.update({
            'ID': f"OPT_{var_id}",
            'Table': 'AppVariables',
            'Column': v.get('Column', v.get('Type', '')),
            'Tags': f"Option , {v.get('Column', v.get('Type', ''))} , ID Connected to Variable",
            'ValueControl': 'Enum',
            'Title': title_en,
            'Description': v.get('Description', ''),
            'UsedFor': f"Option item for {v.get('Column', v.get('Type', ''))}",
            'EnumValue': title_en,
            'Title_hi': title_hi,
            'Title_mr': title_mr,
            'LastEditBy': 'DevNoMi',
            'LastEditOn': now_str
        })
        app_vars.append(row)

    # 3. Question Prompts + Embedded Options (VariableList / EnumList)
    for q in questions:
        q_id = str(q.get('ID', '')).strip()
        if not q_id:
            continue
        
        target_table = q.get('Table', 'Survey')
        target_col = q.get('Column', '')
        opt_vars_raw = str(q.get('OptionVariables', '') or '').strip()
        
        q_title = q.get('Question_en', q.get('Title', ''))
        q_hi = q.get('Question_hi', q.get('Title_hi', ''))
        q_mr = q.get('Question_mr', q.get('Title_mr', ''))

        row = {h: '' for h in STANDARD_HEADERS}
        row.update({
            'ID': f"Q_{q_id}",
            'Table': target_table,
            'Column': target_col,
            'Tags': f"QuestionPrompt , {target_table} , {target_col}",
            'ValueControl': 'VariableList' if opt_vars_raw else 'Enum',
            'Title': q_title,
            'Description': f"Questionnaire code: {q.get('Questionnaire', '')}",
            'UsedFor': f"DisplayName & Options for {target_table}.{target_col}",
            'EnumValue': q_title,
            'VariableList': opt_vars_raw,
            'EnumList': opt_vars_raw,
            'Title_hi': q_hi,
            'Title_mr': q_mr,
            'LastEditBy': 'DevNoMi',
            'LastEditOn': now_str
        })
        app_vars.append(row)

    # 4. Action Grid UI Navigation Controls
    action_modules = [
        ('ACT_Farm', 'Farm', '🚜', 'Farm Details Sub-module'),
        ('ACT_Household', 'Household', '🏠', 'Household Details Sub-module'),
        ('ACT_WaterATM', 'WaterATM', '💧', 'Water & Irrigation Sub-module'),
        ('ACT_Livestock', 'Livestock', '🐄', 'Livestock Sub-module'),
        ('ACT_Collective', 'Collective', '🐷', 'Collectives Sub-module'),
        ('ACT_Efforts', 'Efforts', '⛰️', 'Efforts Sub-module'),
        ('ACT_Complete', 'Complete', '🟢', 'Mark Survey Complete Action')
    ]
    for act_id, name, icon, desc in action_modules:
        row = {h: '' for h in STANDARD_HEADERS}
        row.update({
            'ID': act_id,
            'Table': 'Survey',
            'Column': 'ActionGrid',
            'Tags': 'UI , Navigation , ActionGrid',
            'ValueControl': 'Enum',
            'Title': name,
            'Description': desc,
            'UsedFor': f"Inline Action Grid Button for {name}",
            'EnumValue': name,
            'Title_hi': name,
            'Title_mr': name,
            'ActionIcon': icon,
            'LastEditBy': 'DevNoMi',
            'LastEditOn': now_str
        })
        app_vars.append(row)

    return app_vars


def export_appvariables_csv(app_vars, output_file):
    """
    Exports AppVariables matching the updated Google Sheet headers.
    """
    if not app_vars:
        return
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=STANDARD_HEADERS)
        writer.writeheader()
        writer.writerows(app_vars)


def main():
    parser = argparse.ArgumentParser(description="Ingest questionnaire into updated AppVariables schema.")
    parser.add_argument("xlsx_path", help="Path to client questionnaire Excel file")
    parser.add_argument("--output-dir", "-o", default="./output", help="Directory to save generated output")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    
    print(f"📄 Processing Questionnaire Excel: {args.xlsx_path}")
    questions, variables = parse_questionnaire_excel(args.xlsx_path)
    print(f"   Found {len(questions)} Questions and {len(variables)} Variables.")
    
    app_vars = transform_to_ommnomi_standard_appvariables(questions, variables)
    
    output_csv = os.path.join(args.output_dir, "Generated_AppVariables.csv")
    export_appvariables_csv(app_vars, output_csv)
    print(f"✅ Generated AppVariables CSV matching updated schema: {output_csv}")

if __name__ == "__main__":
    main()
