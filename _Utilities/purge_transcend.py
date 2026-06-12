import openpyxl

wb = openpyxl.load_workbook('/Users/ommnomi/AppSheets/Projects/Transcend/Transcend Main.xlsx')
ws = wb['FormIntake']

cols_to_delete = [
    'Card Number', 
    'Card Expiration Date', 
    'Security Code',
    'Medication Name',
    'Dosage',
    'Form / Unit (Capsule, tablet, liquid)',
    'Amount (How many are taken)',
    'Frequency (How often is it taken)',
    'Symptoms Being Treated',
    'Prescribed By ',
    'Is there another medication that you take? '
]

# We must delete from right to left to avoid index shifting
indices_to_delete = []

for idx, cell in enumerate(ws[1]):
    if cell.value and cell.value.strip() in [c.strip() for c in cols_to_delete]:
        indices_to_delete.append(idx + 1)

print(f"Found {len(indices_to_delete)} columns to delete.")

for col_idx in sorted(indices_to_delete, reverse=True):
    header = ws.cell(row=1, column=col_idx).value
    print(f"Deleting column {col_idx}: {header}")
    ws.delete_cols(col_idx)

wb.save('/Users/ommnomi/AppSheets/Projects/Transcend/Transcend Main.xlsx')
print("Successfully saved Transcend Main.xlsx.")
