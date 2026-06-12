#!/usr/bin/env python3
"""
AppSheet Documentation Parser — OmmNoMi Full Schema Engine (V4)
================================================================
Parses the standard AppSheet "Application Documentation" HTML export
and produces a clean, AI-context-optimized Markdown schema.

Optionally cross-references an XLSX spreadsheet export to determine
physical vs virtual columns with 100% accuracy.

Usage:
    python3 parse_appdoc.py <path_to_html> --compact
    python3 parse_appdoc.py <path_to_html> --compact --spreadsheet <path_to_xlsx>
    python3 parse_appdoc.py <path_to_html> --output <path_to_md>

If --output is omitted, the markdown is written next to the HTML file
as {AppName}_schema.md (compact) or {AppName}_parsed.md (verbose).
"""

import re
import sys
import os
import json
import html
import argparse
from collections import OrderedDict


# ─── HTML Helpers ────────────────────────────────────────────────────────────

def strip_tags(text: str) -> str:
    """Remove all HTML tags and decode entities."""
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    return text.strip()


def extract_td_pairs(html_block: str) -> OrderedDict:
    """Extract label → value pairs from <tr><td>label</td><td>value</td></tr> rows."""
    pairs = OrderedDict()
    rows = re.findall(r'<tr>(.*?)</tr>', html_block, re.DOTALL)
    for row in rows:
        tds = re.findall(r'<td>(.*?)</td>', row, re.DOTALL)
        if len(tds) >= 2:
            label = strip_tags(tds[0])
            value = strip_tags(tds[1])
            if label and value:
                pairs[label] = value
    return pairs


# ─── Section Extractors ─────────────────────────────────────────────────────

def extract_app_info(content: str) -> dict:
    """Extract top-level app metadata."""
    info = {}

    # App name
    m = re.search(r'<h2>(.*?)</h2>', content[600:800], re.DOTALL)
    if m:
        info['App Name'] = strip_tags(m.group(1))

    # Generated date
    m = re.search(r'id="generatedDate">(.*?)</span>', content)
    if m:
        info['Generated At'] = strip_tags(m.group(1))

    # Summary counts
    m = re.search(r'(\d+)\s*Tables', content)
    if m:
        info['Tables'] = m.group(1)
    m = re.search(r'(\d+)\s*Columns', content)
    if m:
        info['Columns'] = m.group(1)
    m = re.search(r'(\d+)\s*Slices', content)
    if m:
        info['Slices'] = m.group(1)
    m = re.search(r'(\d+)\s*Views', content)
    if m:
        info['Views'] = m.group(1)
    m = re.search(r'(\d+)\s*Actions', content)
    if m:
        info['Actions'] = m.group(1)
    m = re.search(r'(\d+)\s*Format Rules', content)
    if m:
        info['Format Rules'] = m.group(1)
    m = re.search(r'(\d+)\s*Workflow Rules', content)
    if m:
        info['Workflow Rules'] = m.group(1)

    # Version, Short Name, etc.
    for field in ['Short Name', 'Short Description', 'Version', 'Default app folder', 'Function']:
        label_escaped = re.escape(field)
        m = re.search(
            rf'title="[^"]*">{label_escaped}</label>\s*</td>\s*<td>(.*?)</td>',
            content, re.DOTALL
        )
        if m:
            info[field] = strip_tags(m.group(1))

    # Runnable / Deployable
    m = re.search(r'Runnable\?</label></td><td>(.*?)</td>', content)
    if m:
        info['Runnable'] = strip_tags(m.group(1))
    m = re.search(r'Deployable\?</label></td><td>(.*?)</td>', content)
    if m:
        info['Deployable'] = strip_tags(m.group(1))

    return info


def extract_tables(content: str) -> list:
    """Extract table definitions (non-schema tables only)."""
    tables = []

    # Find all h5 IDs for tables
    all_ids = re.findall(r'<h5 id="table_([^"]+)">', content)
    table_ids = [t for t in all_ids if not t.endswith('_Schema')]

    for table_id in table_ids:
        # Extract the block between this h5 and the next h5
        escaped_id = re.escape(table_id)
        pattern = rf'<h5 id="table_{escaped_id}">(.*?)(?=<h5 id="table_|<div class="flexContainer pageBreak">)'
        m = re.search(pattern, content, re.DOTALL)
        if m:
            block = m.group(1)
            props = extract_td_pairs(block)
            tables.append({
                'name': table_id,
                'properties': props
            })

    return tables


def extract_columns_for_schema(content: str, schema_name: str) -> list:
    """Extract columns from a schema section."""
    columns = []

    # Find the schema section
    pattern = rf'<h5 id="table_{re.escape(schema_name)}_Schema">(.*?)(?=<h5 id="table_|<div class="flexContainer pageBreak">)'
    m = re.search(pattern, content, re.DOTALL)
    if not m:
        return columns

    schema_block = m.group(1)

    # Find each column: <h3 id="table_XXX_Schema_colN">Column N: ColumnName</h3>
    col_pattern = r'<h3[^>]*>Column \d+:\s*(.*?)</h3>(.*?)(?=<h3[^>]*>Column \d+:|$)'
    col_matches = re.finditer(col_pattern, schema_block, re.DOTALL)

    for cm in col_matches:
        col_name = strip_tags(cm.group(1))
        col_block = cm.group(2)
        props = extract_td_pairs(col_block)

        # Parse TypeAuxData JSON if present
        type_details = {}
        if 'Type Qualifier' in props:
            try:
                # AppSheet sometimes escapes quotes in the HTML export
                qual_str = html.unescape(props['Type Qualifier'])
                type_details = json.loads(qual_str)
                # Clean out null values
                type_details = {k: v for k, v in type_details.items() if v is not None and v != ''}
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Extract advanced logic if available
        ref_table = type_details.get('Ref_Table_Name', '')
        show_if = type_details.get('Show_If', '')
        valid_if = type_details.get('Valid_If', '')
        required_if = type_details.get('Required_If', '')
        editable_if = type_details.get('Editable_If', '')
        
        # Extract Enum values
        enum_values = type_details.get('EnumValues', '')
        if not enum_values:
            # Sometimes stored as a ValidIf list for enums
            pass

        columns.append({
            'name': col_name,
            'type': props.get('Type', ''),
            'key': props.get('Is Key', 'No'),
            'label': props.get('Is Label', 'No'),
            'hidden': props.get('Hidden', 'No'),
            'read_only': props.get('Read-Only', 'No'),
            'required': props.get('Required', 'No'),
            'initial_value': props.get('App Formula', props.get('Initial value', '')),
            'description': props.get('Description', ''),
            'display_name': props.get('Display name', ''),
            'ref_table': ref_table,
            'show_if': show_if,
            'valid_if': valid_if,
            'required_if': required_if,
            'editable_if': editable_if,
            'enum_values': enum_values,
            'type_details': type_details,
            'all_props': props
        })

    return columns


def extract_slices(content: str) -> list:
    """Extract slice definitions."""
    slices = []
    pattern = r'<h5 id="slice_([^"]+)">(.*?)(?=<h5 id="slice_|<div class="flexContainer pageBreak">|<h4 class="flex1 subtitle")'
    matches = re.finditer(pattern, content, re.DOTALL)

    for m in matches:
        slice_name = m.group(1)
        block = m.group(2)
        props = extract_td_pairs(block)
        slices.append({
            'name': slice_name,
            'properties': props
        })

    return slices


def extract_views(content: str) -> list:
    """Extract view definitions from the UX section."""
    views = []

    # Views are in the UX/Views section. Pattern: <h5 id="view_XXX">
    pattern = r'<h5 id="view_([^"]+)">(.*?)(?=<h5 id="view_|<div class="flexContainer pageBreak">|<h4 class="flex1 subtitle")'
    matches = re.finditer(pattern, content, re.DOTALL)

    for m in matches:
        view_name = m.group(1)
        block = m.group(2)
        props = extract_td_pairs(block)
        views.append({
            'name': view_name,
            'properties': props
        })

    # Fallback: find all View name entries
    if not views:
        views = _extract_views_robust(content)

    return views


def _extract_views_robust(content: str) -> list:
    """More robust view extraction by finding all View name entries."""
    views = []
    pattern = r'View name</label>\s*</td>\s*<td>(.*?)</td>'
    names = re.findall(pattern, content, re.DOTALL)

    for name_html in names:
        name = strip_tags(name_html)
        if name and not name.startswith('_'):
            escaped_name = re.escape(name_html)
            block_match = re.search(
                rf'View name</label>\s*</td>\s*<td>{escaped_name}</td>(.*?)(?=View name</label>\s*</td>|</section>)',
                content, re.DOTALL
            )
            props = {}
            if block_match:
                props = extract_td_pairs(block_match.group(1))
            props['View name'] = name

            views.append({
                'name': name,
                'properties': props
            })

    return views


def extract_actions(content: str) -> list:
    """Extract action definitions from the Behavior section."""
    actions = []

    # Find the action section
    action_section_match = re.search(r'class="actionSection">(.*?)(?=<div class="flexContainer pageBreak">|class="automationSection">)', content, re.DOTALL)
    if not action_section_match:
        return _extract_actions_robust(content)

    action_section = action_section_match.group(1)

    # Split by action blocks
    action_blocks = re.split(r'(?=<tr>\s*<td>\s*<label[^>]*>Action name</label>\s*</td>)', action_section)

    for block in action_blocks:
        if not block.strip():
            continue
        props = extract_td_pairs(block)
        if 'Action name' in props or 'Action name' in block:
            name = props.get('Action name', '')
            if not name:
                m = re.search(r'Action name\s*</td>\s*<td>(.*?)</td>', block, re.DOTALL)
                if m:
                    name = strip_tags(m.group(1))

            if not name:
                m = re.search(r'>\s*Action name\s*</\w+>\s*</\w+>\s*<\w+[^>]*>\s*<[^>]*>\s*<[^>]*>\s*(.*?)\s*<', block, re.DOTALL)
                if m:
                    name = strip_tags(m.group(1))

            if name:
                actions.append({
                    'name': name,
                    'table': props.get('For a record of this table', ''),
                    'do_this': props.get('Do this', ''),
                    'only_if': props.get('Only if this condition is true', ''),
                    'prominence': props.get('Prominence', ''),
                    'visible': props.get('Visible?', ''),
                    'modifies_data': props.get('Modifies data?', ''),
                    'display_name': props.get('Display name', ''),
                    'all_props': props
                })

    if not actions:
        return _extract_actions_robust(content)
    return actions


def _extract_actions_robust(content: str) -> list:
    """More robust action extraction."""
    actions = []
    pattern = r'Action name</label>\s*</td>\s*<td>(.*?)</td>'
    names = re.findall(pattern, content, re.DOTALL)

    for name_html in names:
        name = strip_tags(name_html)
        if not name:
            continue

        escaped_name = re.escape(name_html)
        block_match = re.search(
            rf'Action name</label>\s*</td>\s*<td>{escaped_name}</td>(.*?)(?=Action name</label>\s*</td>|</section>)',
            content, re.DOTALL
        )
        props = {}
        if block_match:
            props = extract_td_pairs(block_match.group(1))

        actions.append({
            'name': name,
            'table': props.get('For a record of this table', ''),
            'do_this': props.get('Do this', ''),
            'only_if': props.get('Only if this condition is true', ''),
            'prominence': props.get('Prominence', ''),
            'visible': props.get('Visible?', ''),
            'modifies_data': props.get('Modifies data?', ''),
            'display_name': props.get('Display name', ''),
            'all_props': props
        })

    return actions


def extract_bots(content: str) -> list:
    """Extract Bot definitions."""
    bots = []
    
    pattern = r'Bot name</label>\s*</td>\s*<td>(.*?)</td>'
    names = re.findall(pattern, content, re.DOTALL)
    
    seen_bots = set()

    for name_html in names:
        name = strip_tags(name_html)
        if not name or name in seen_bots:
            continue
            
        seen_bots.add(name)

        escaped_name = re.escape(name_html)
        block_match = re.search(
            rf'Bot name</label>\s*</td>\s*<td>{escaped_name}</td>(.*?)(?=Bot name</label>\s*</td>|</section>|<div class="flexContainer pageBreak">)',
            content, re.DOTALL
        )
        props = {}
        if block_match:
            props = extract_td_pairs(block_match.group(1))

        event_name = props.get('Event name', '')
        event_type = props.get('Event type', '')
        event_table = props.get('Table', '')
        event_condition = props.get('Condition', '')
        
        bots.append({
            'name': name,
            'event': event_name,
            'event_type': event_type,
            'table': event_table,
            'condition': event_condition,
            'all_props': props
        })

    return bots


# ─── XLSX Spreadsheet Reader ────────────────────────────────────────────────

def read_spreadsheet_headers(xlsx_path: str) -> dict:
    """Read worksheet names and column headers from an XLSX file.
    
    Returns:
        dict mapping worksheet_name -> list of column header strings
    """
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl not installed. Run: pip install openpyxl")
        print("   Skipping spreadsheet cross-reference.")
        return {}
    
    headers = {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            row1 = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
                if cell is not None:
                    row1.append(str(cell).strip())
            headers[ws_name] = row1
        wb.close()
    except Exception as e:
        print(f"⚠️  Error reading spreadsheet: {e}")
    
    return headers


def match_tables_to_worksheets(tables: list, ws_headers: dict) -> dict:
    """Match AppSheet table names to worksheet names.
    
    Returns dict: table_name -> {'worksheet': ws_name, 'physical_cols': [col_names]}
    """
    mapping = {}
    
    # Build a lookup: normalized worksheet name -> original name
    ws_lookup = {}
    for ws_name in ws_headers:
        ws_lookup[ws_name.lower().replace(' ', '').replace('_', '')] = ws_name
    
    for t in tables:
        table_name = t['name']
        ws_name_from_props = t['properties'].get('Worksheet Name/Qualifier', '')
        
        matched_ws = None
        
        # Try exact match from properties first
        if ws_name_from_props and ws_name_from_props in ws_headers:
            matched_ws = ws_name_from_props
        else:
            # Try normalized match
            norm_table = table_name.lower().replace(' ', '').replace('_', '')
            if norm_table in ws_lookup:
                matched_ws = ws_lookup[norm_table]
            else:
                # Try partial match
                for ws_key, ws_val in ws_lookup.items():
                    if norm_table in ws_key or ws_key in norm_table:
                        matched_ws = ws_val
                        break
        
        if matched_ws:
            mapping[table_name] = {
                'worksheet': matched_ws,
                'physical_cols': ws_headers[matched_ws]
            }
    
    return mapping


# ─── Formula Formatter ───────────────────────────────────────────────────────

# Functions that warrant indented multi-line formatting
_MAJOR_FUNCS = frozenset([
    'IF', 'IFS', 'SELECT', 'FILTER', 'AND', 'OR', 'ANY',
    'LOOKUP', 'REF_ROWS', 'SORT', 'SWITCH', 'IN', 'CONTAINS',
    'ORDERBY', 'MAXROW', 'MINROW', 'INDEX', 'ISNOTBLANK',
])


def fmt(expr: str) -> str:
    """Format an AppSheet expression for clean AI consumption.
    
    Rules:
    - If expression is empty or has no function calls → return as-is
    - If expression is very short (≤40 chars) → return as-is (inline)
    - Otherwise → pretty-print with indentation
    """
    if not expr:
        return ''
    
    # Strip any pre-existing artificial newlines from HTML export
    clean = expr.replace('\n', ' ')
    clean = re.sub(r'\s+', ' ', clean).strip()
    
    # Short expressions stay inline
    if len(clean) <= 40 or '(' not in clean:
        return clean
    
    # For expressions starting with =, format the body
    if clean.startswith('='):
        body = clean[1:].strip()
        formatted = _indent_expr(body)
        return f"={formatted}"
    else:
        return _indent_expr(clean)


def _indent_expr(expr: str) -> str:
    """Indent an AppSheet expression using parenthesis-level tracking."""
    output = []
    level = 0
    in_quote = False
    i = 0
    stack = []
    
    while i < len(expr):
        c = expr[i]
        
        if c == '"':
            in_quote = not in_quote
            output.append(c)
        elif not in_quote:
            if c == '(':
                # Lookback to find the function name
                func_name = _lookback_func(output)
                output.append(c)
                if func_name.upper() in _MAJOR_FUNCS:
                    level += 1
                    stack.append(True)
                    output.append('\n' + '  ' * level)
                else:
                    stack.append(False)
            elif c == ')':
                # If we indented into this level, outdent before closing
                indented = stack.pop() if stack else False
                if indented:
                    level -= 1
                    output.append('\n' + '  ' * level)
                output.append(c)
            elif c == ',':
                output.append(c)
                if level > 0:
                    output.append('\n' + '  ' * level)
            else:
                output.append(c)
        else:
            output.append(c)
        i += 1
    
    result = ''.join(output)
    # Clean trailing whitespace on lines
    result = '\n'.join(line.rstrip() for line in result.split('\n'))
    # Remove empty lines
    result = re.sub(r'\n\s*\n', '\n', result)
    return result


def _lookback_func(output: list) -> str:
    """Look backwards through output chars to find a function name before '('."""
    word = []
    idx = len(output) - 1
    while idx >= 0:
        c = output[idx]
        if c.isalpha() or c == '_':
            word.append(c)
            idx -= 1
        else:
            break
    return ''.join(reversed(word))


def _lookback_func_for_close(output: list) -> str:
    """Check if the current closing paren corresponds to a major function indent.
    Returns a truthy string if we should outdent."""
    # Simple heuristic: if recent output contains a newline with indentation,
    # we probably indented this block
    recent = ''.join(output[-20:]) if len(output) >= 20 else ''.join(output)
    return '\n' in recent


# ─── Compact Markdown Generator (AI-context optimized) ──────────────────────

def generate_compact_markdown(app_info, tables, all_columns, slices, views, actions, bots, table_ws_map=None) -> str:
    """Generate a dense, AI-friendly Markdown optimized for context windows."""
    lines = []

    app_name = app_info.get('App Name', app_info.get('Short Name', 'AppSheet App'))
    version = app_info.get('Version', '?')
    generated = app_info.get('Generated At', '?')

    lines.append(f"# {app_name} — AppSheet Schema (v{version})")
    lines.append(f"> Parsed: {generated} | {app_info.get('Tables','?')}T / {app_info.get('Columns','?')}C / {app_info.get('Slices','?')}S / {app_info.get('Views','?')}V / {app_info.get('Actions','?')}A / {app_info.get('Format Rules','?')}FR")
    lines.append(f"> Deployable: {app_info.get('Deployable','?')} | Runnable: {app_info.get('Runnable','?')}")
    if table_ws_map:
        lines.append(f"> Spreadsheet cross-reference: {len(table_ws_map)} tables matched to worksheets")
    lines.append("")

    # ─── Tables (one-line each) ───
    lines.append("## Tables")
    lines.append("```")
    for t in tables:
        p = t['properties']
        source = p.get('Data Source', '?')
        worksheet = p.get('Worksheet Name/Qualifier', '?')
        updates = p.get('Are updates allowed?', '?')
        ws_info = ""
        if table_ws_map and t['name'] in table_ws_map:
            ws_info = f"  [XLSX: {table_ws_map[t['name']]['worksheet']}]"
        lines.append(f"  {t['name']:25s} src={source:8s} sheet={worksheet:20s} mode={updates}{ws_info}")
    lines.append("```")
    lines.append("")

    # ─── Columns (with Schema Deduplication) ───
    lines.append("## Columns")
    
    # Build slice map for cross-referencing
    slice_map = {}
    if slices:
        for s in slices:
            source = s['properties'].get('Source Table', s['properties'].get('Source table', '?'))
            slice_map[s['name']] = source

    # Identify all tables as potential masters, sorted by size descending (largest first)
    master_tables = [t['name'] for t in tables]
    master_tables.sort(key=lambda name: len(all_columns.get(name, [])), reverse=True)
    
    # Build physical column sets from spreadsheet if available
    physical_cols_map = {}
    if table_ws_map:
        for tname, info in table_ws_map.items():
            physical_cols_map[tname] = set(info['physical_cols'])
            
    for table_name, columns in all_columns.items():
        if not columns:
            continue
        
        # ─── Schema Deduplication ───
        matched_master = None
        additions = []
        
        if len(columns) > 10:
            my_col_names = {c['name'] for c in columns}
            for mt in master_tables:
                if mt == table_name:
                    continue  # Don't inherit from self
                    
                mt_cols = all_columns.get(mt, [])
                if not mt_cols or len(mt_cols) < 5:
                    continue
                mt_col_names = {c['name'] for c in mt_cols}
                
                # Match if ≥90% of master columns exist in this table
                overlap = mt_col_names & my_col_names
                if len(overlap) >= len(mt_col_names) * 0.90:
                    matched_master = mt
                    additions = [c['name'] for c in columns if c['name'] not in mt_col_names]
                    break
        
        lines.append(f"### {table_name} ({len(columns)} cols)")
        
        if matched_master:
            mt_len = len(all_columns.get(matched_master, []))
            lines.append(f"[Inherits all {mt_len} columns from Table: {matched_master}]")
            if additions:
                add_strs = []
                for a_name in additions:
                    a_col = next((c for c in columns if c['name'] == a_name), None)
                    if a_col:
                        add_strs.append(f"{a_name}: {a_col['type']}")
                    else:
                        add_strs.append(a_name)
                lines.append(f"+ Modified/Added Columns:")
                for s in add_strs:
                    lines.append(f"  - {s}")
            lines.append("")
            continue

        # ─── Full Column Listing ───
        physical_set = physical_cols_map.get(table_name)
        
        lines.append("```")
        for col in columns:
            col_name = col['name']
            flags = []
            
            # System column
            if col_name == '_RowNumber':
                flags.append("SYSTEM")
            
            # Primary Key
            if col['key'] == 'Yes':
                flags.append("PK")
            
            # Label
            if col['label'] == 'Yes':
                flags.append("LABEL")
            
            # Hidden
            if col['hidden'] == 'Yes':
                flags.append("HIDDEN")
            
            # Read-Only
            if col['read_only'] == 'Yes':
                flags.append("RO")
            
            # Virtual Column detection
            is_vc = False
            if physical_set is not None:
                # Absolute detection via spreadsheet cross-reference
                if col_name not in physical_set and col_name != '_RowNumber':
                    is_vc = True
            else:
                # Heuristic detection
                if col_name.startswith('_Computed') or col_name.startswith('Related '):
                    is_vc = True
                elif col['initial_value'] and 'REF_ROWS' in col['initial_value'].upper():
                    is_vc = True
                elif 'App Formula' in col['all_props'] and col['read_only'] == 'Yes' and col_name != '_RowNumber':
                    is_vc = True
            
            if is_vc:
                flags.append("VC")
            
            flag_str = f" [{','.join(flags)}]" if flags else ""
            
            # Type and Ref Table
            type_str = col['type']
            if col['ref_table']:
                type_str = f"Ref -> Table: {col['ref_table']}"
            
            # Enum values
            enum_str = ""
            if col['enum_values']:
                vals = col['enum_values']
                if isinstance(vals, list):
                    enum_str = f" [Values: {', '.join(repr(v) for v in vals[:8])}]"
                    if len(vals) > 8:
                        enum_str = enum_str[:-1] + f", ... +{len(vals)-8} more]"
                elif isinstance(vals, str) and vals:
                    enum_str = f" [Values: {vals}]"
            
            # Initial value / App Formula (formatted)
            init = col['initial_value']
            init_str = f" = {fmt(init)}" if init else ""
            
            # Display name
            display = f' (→"{col["display_name"]}")' if col['display_name'] else ""
            
            # Dynamic Logic
            logic_flags = []
            if col['show_if']:
                logic_flags.append(f"ShowIf: `{fmt(col['show_if'])}`")
            if col['valid_if']:
                logic_flags.append(f"ValidIf: `{fmt(col['valid_if'])}`")
            if col['required_if']:
                logic_flags.append(f"ReqIf: `{fmt(col['required_if'])}`")
            if col['editable_if']:
                logic_flags.append(f"EditIf: `{fmt(col['editable_if'])}`")
            
            logic_str = f" {{ {' | '.join(logic_flags)} }}" if logic_flags else ""
            
            # Check for slice cross-references in formulas
            combined_formulas = init_str + logic_str
            found_slices = []
            if combined_formulas:
                for sname, ssource in slice_map.items():
                    # Match SliceName[ or SliceName( or SliceName,
                    if re.search(r'\b' + re.escape(sname) + r'\b(?:\[|\s*[,)])', combined_formulas):
                        found_slices.append(f"{sname} -> {ssource}")
            
            slice_cross_ref = f" {{ Slices Cross-Ref: {', '.join(found_slices)} }}" if found_slices else ""

            lines.append(f"  {col_name}: {type_str}{flag_str}{enum_str}{init_str}{display}{logic_str}{slice_cross_ref}")
        lines.append("```")
        lines.append("")

    # ─── Slices ───
    if slices:
        lines.append("## Slices")
        for s in slices:
            p = s['properties']
            source = p.get('Source Table', p.get('Source table', '?'))
            row_filter = p.get('Row filter condition', p.get('Row Filter', ''))
            lines.append(f"- **{s['name']}** ({source}): `{row_filter}`")
        lines.append("")

    # ─── Views (compact, skip auto-generated _Detail/_Form) ───
    if views:
        lines.append("## Views")
        
        custom_views = []
        auto_views = []
        for v in views:
            name = v['name']
            if name.endswith('_Detail') or name.endswith('_Form'):
                auto_views.append(name)
            else:
                custom_views.append(v)

        if custom_views:
            lines.append("### Custom Views")
            for v in custom_views:
                p = v['properties']
                vtype = p.get('View type', '?')
                data = p.get('For this data', '?')
                position = p.get('Position', '')
                pos_str = f" pos={position}" if position else ""
                lines.append(f"- **{v['name']}**: {vtype} → {data}{pos_str}")
        
        if auto_views:
            lines.append(f"### Auto-generated ({len(auto_views)})")
            lines.append(f"  {', '.join(auto_views)}")
        lines.append("")

    # ─── Actions (compact, deduplicated, skip standard CRUD) ───
    if actions:
        lines.append("## Actions")

        actions_by_table = OrderedDict()
        for a in actions:
            table = a['table'] or 'Unknown'
            if table not in actions_by_table:
                actions_by_table[table] = []
            actions_by_table[table].append(a)

        for table, table_actions in actions_by_table.items():
            standard_types = {'DELETE_RECORD', 'EDIT_RECORD', 'ADD_RECORD', 'CALL', 'SMS', 'EMAIL',
                              'NAVIGATE_APP', 'NAVIGATE_URL', 'OPEN_FILE'}

            custom = []
            auto_count = 0
            auto_types = set()
            for a in table_actions:
                if a['do_this'] in standard_types and not a['display_name']:
                    auto_count += 1
                    auto_types.add(a['do_this'])
                else:
                    custom.append(a)

            seen = set()
            deduped_custom = []
            for a in custom:
                if a['name'] not in seen:
                    seen.add(a['name'])
                    deduped_custom.append(a)

            lines.append(f"### {table}")
            if auto_count > 0:
                lines.append(f"  _Auto ({auto_count}): {', '.join(sorted(auto_types))}_")

            for a in deduped_custom:
                only_if = a['only_if'][:80] if a['only_if'] else ''
                display = f" (→\"{a['display_name']}\")" if a['display_name'] else ""
                if_str = f" IF `{only_if}`" if only_if else ""
                lines.append(f"  - **{a['name']}{display}**: {a['do_this']}{if_str}")

            lines.append("")

    # ─── Bots / Automations ───
    if bots:
        lines.append("## Bots & Automations")
        for b in bots:
            table_str = f" [{b['table']}]" if b['table'] else ""
            event_str = f" ({b['event_type']} '{b['event']}')" if b['event_type'] else ""
            cond_str = f" IF `{b['condition']}`" if b['condition'] else ""
            lines.append(f"- **{b['name']}**{table_str}{event_str}{cond_str}")
        lines.append("")

    # ─── Observations ───
    lines.append("## Observations")
    
    # Fragile keys
    for table_name, columns in all_columns.items():
        if table_name.startswith('_') or 'Output' in table_name or 'Process Table' in table_name:
            continue
        for col in columns:
            if col['name'] == '_RowNumber' and col['key'] == 'Yes':
                lines.append(f"- ⚠️ **{table_name}** uses `_RowNumber` as key (fragile)")

    # Missing labels
    for table_name, columns in all_columns.items():
        if table_name.startswith('_') or 'Output' in table_name or 'Process Table' in table_name:
            continue
        has_label = any(col['label'] == 'Yes' for col in columns)
        if not has_label:
            lines.append(f"- ℹ️ **{table_name}** has no Label column")

    lines.append("")

    return '\n'.join(lines)


# ─── Verbose Markdown Generator ─────────────────────────────────────────────

def generate_markdown(app_info, tables, all_columns, slices, views, actions) -> str:
    """Generate clean, organized Markdown from parsed data (verbose format)."""
    lines = []

    # ─── Header ───
    app_name = app_info.get('App Name', app_info.get('Short Name', 'AppSheet App'))
    lines.append(f"# {app_name} — AppSheet Configuration Report")
    lines.append(f"> Auto-generated from AppSheet Documentation HTML export")
    lines.append(f"> Generated: {app_info.get('Generated At', 'Unknown')}")
    lines.append("")

    # ─── App Summary ───
    lines.append("## App Summary")
    lines.append("")
    lines.append("| Property | Value |")
    lines.append("|----------|-------|")
    for k, v in app_info.items():
        lines.append(f"| {k} | {v} |")
    lines.append("")

    # ─── Tables ───
    lines.append("---")
    lines.append("")
    lines.append("## Tables")
    lines.append("")

    lines.append("| # | Table Name | Source | Worksheet | Updates Allowed |")
    lines.append("|---|-----------|--------|-----------|-----------------|")
    for i, t in enumerate(tables, 1):
        p = t['properties']
        source = p.get('Data Source', '')
        worksheet = p.get('Worksheet Name/Qualifier', p.get('Source Path', ''))
        updates = p.get('Are updates allowed?', '')
        lines.append(f"| {i} | **{t['name']}** | {source} | {worksheet} | {updates} |")
    lines.append("")

    # ─── Column Details Per Table ───
    lines.append("---")
    lines.append("")
    lines.append("## Column Details")
    lines.append("")

    for table_name, columns in all_columns.items():
        if not columns:
            continue
        if table_name.startswith('_'):
            continue

        lines.append(f"### {table_name}")
        lines.append("")
        lines.append("| Column | Type | Key | Label | Hidden | Read-Only | Initial Value / App Formula |")
        lines.append("|--------|------|-----|-------|--------|-----------|---------------------------|")

        for col in columns:
            if col['name'] == '_RowNumber':
                continue
            key = '🔑' if col['key'] == 'Yes' else ''
            label = '🏷️' if col['label'] == 'Yes' else ''
            hidden = '👁️‍🗨️' if col['hidden'] == 'Yes' else ''
            ro = '🔒' if col['read_only'] == 'Yes' else ''
            init = col['initial_value'][:60] if col['initial_value'] else ''
            display = f" ({col['display_name']})" if col['display_name'] else ''
            lines.append(f"| {col['name']}{display} | {col['type']} | {key} | {label} | {hidden} | {ro} | `{init}` |")

        lines.append("")

    # ─── Slices ───
    if slices:
        lines.append("---")
        lines.append("")
        lines.append("## Slices")
        lines.append("")
        lines.append("| Slice Name | Source Table | Row Filter | Description |")
        lines.append("|-----------|-------------|-----------|-------------|")
        for s in slices:
            p = s['properties']
            source = p.get('Source Table', p.get('Source table', ''))
            row_filter = p.get('Row filter condition', p.get('Row Filter', ''))
            desc = p.get('Description', p.get('Slice description', ''))
            lines.append(f"| **{s['name']}** | {source} | `{row_filter}` | {desc} |")
        lines.append("")

    # ─── Views ───
    if views:
        lines.append("---")
        lines.append("")
        lines.append("## Views")
        lines.append("")
        lines.append("| View Name | Type | Data Source | Position | Sort By |")
        lines.append("|-----------|------|-----------|----------|---------|")
        for v in views:
            p = v['properties']
            vtype = p.get('View type', '')
            data = p.get('For this data', '')
            position = p.get('Position', '')
            sort = p.get('Sort by', '')
            lines.append(f"| **{v['name']}** | {vtype} | {data} | {position} | {sort} |")
        lines.append("")

    # ─── Actions ───
    if actions:
        lines.append("---")
        lines.append("")
        lines.append("## Actions")
        lines.append("")

        actions_by_table = OrderedDict()
        for a in actions:
            table = a['table'] or 'Unknown'
            if table not in actions_by_table:
                actions_by_table[table] = []
            actions_by_table[table].append(a)

        for table, table_actions in actions_by_table.items():
            lines.append(f"### {table}")
            lines.append("")
            lines.append("| Action | Do This | Only If | Prominence | Visible |")
            lines.append("|--------|---------|---------|-----------|---------|")
            for a in table_actions:
                only_if = a['only_if'][:50] if a['only_if'] else ''
                display = f" ({a['display_name']})" if a['display_name'] else ''
                lines.append(f"| {a['name']}{display} | {a['do_this']} | `{only_if}` | {a['prominence']} | {a['visible']} |")
            lines.append("")

    # ─── Observations ───
    lines.append("---")
    lines.append("")
    lines.append("## Observations & Potential Issues")
    lines.append("")

    for table_name, columns in all_columns.items():
        if table_name.startswith('_'):
            continue
        for col in columns:
            if col['name'] == '_RowNumber' and col['key'] == 'Yes':
                lines.append(f"> [!WARNING]")
                lines.append(f"> **{table_name}** uses `_RowNumber` as its key. This is fragile and should be replaced with `TEXT(UNIQUEID())`.")
                lines.append("")

    for table_name, columns in all_columns.items():
        if table_name.startswith('_'):
            continue
        has_label = any(col['label'] == 'Yes' for col in columns)
        if not has_label:
            lines.append(f"> [!NOTE]")
            lines.append(f"> **{table_name}** has no column marked as Label. Consider adding a Label virtual column.")
            lines.append("")

    return '\n'.join(lines)


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='AppSheet Documentation Parser — OmmNoMi Full Schema Engine (V4)',
        epilog='Example: python3 parse_appdoc.py app_doc.html --compact --spreadsheet database.xlsx'
    )
    parser.add_argument('html_file', help='Path to the AppSheet Documentation HTML file')
    parser.add_argument('--output', '-o', help='Output markdown file path (default: {AppName}_schema.md)')
    parser.add_argument('--compact', '-c', action='store_true',
                        help='Generate compact AI-context-optimized output (default: verbose)')
    parser.add_argument('--spreadsheet', '-s', nargs='+',
                        help='Paths to one or more XLSX/CSV spreadsheets for physical column cross-reference')
    args = parser.parse_args()

    html_path = args.html_file
    if not os.path.exists(html_path):
        print(f"Error: File not found: {html_path}")
        sys.exit(1)

    print(f"📄 Reading: {html_path}")
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    print("🔍 Extracting app info...")
    app_info = extract_app_info(content)
    app_name = app_info.get('App Name', app_info.get('Short Name', 'AppSheet'))
    # Sanitize app name for filename
    safe_name = re.sub(r'[^\w\s-]', '', app_name).strip().replace(' ', '_')

    print("📊 Extracting tables...")
    tables = extract_tables(content)
    print(f"   Found {len(tables)} tables")

    print("📋 Extracting columns...")
    all_columns = OrderedDict()
    for t in tables:
        cols = extract_columns_for_schema(content, t['name'])
        all_columns[t['name']] = cols
        if cols:
            print(f"   {t['name']}: {len(cols)} columns")

    print("🔪 Extracting slices...")
    slices = extract_slices(content)
    print(f"   Found {len(slices)} slices")

    print("👁️ Extracting views...")
    views = extract_views(content)
    print(f"   Found {len(views)} views")

    print("⚡ Extracting actions...")
    actions = extract_actions(content)
    print(f"   Found {len(actions)} actions")

    print("🤖 Extracting bots...")
    bots = extract_bots(content)
    print(f"   Found {len(bots)} bots")

    # ─── Spreadsheet Cross-Reference ───
    table_ws_map = None
    if args.spreadsheet:
        all_ws_headers = {}
        for xlsx_path in args.spreadsheet:
            if not os.path.exists(xlsx_path):
                print(f"⚠️  Spreadsheet not found: {xlsx_path}")
            else:
                print(f"📑 Reading spreadsheet: {xlsx_path}")
                ws_headers = read_spreadsheet_headers(xlsx_path)
                if ws_headers:
                    all_ws_headers.update(ws_headers)
        
        if all_ws_headers:
            table_ws_map = match_tables_to_worksheets(tables, all_ws_headers)
            print(f"   Matched {len(table_ws_map)}/{len(tables)} tables to worksheets")
            for tname, info in table_ws_map.items():
                print(f"   {tname} → {info['worksheet']} ({len(info['physical_cols'])} physical cols)")

    # ─── Output ───
    if args.output:
        output_path = args.output
    else:
        output_dir = os.path.dirname(html_path)
        suffix = '_schema.md' if args.compact else '_parsed.md'
        output_path = os.path.join(output_dir, safe_name + suffix)

    print(f"📝 Generating {'compact' if args.compact else 'verbose'} markdown...")
    if args.compact:
        markdown = generate_compact_markdown(app_info, tables, all_columns, slices, views, actions, bots, table_ws_map)
    else:
        # NOTE: Verbose doesn't support bots or spreadsheet cross-ref yet
        markdown = generate_markdown(app_info, tables, all_columns, slices, views, actions)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(markdown)

    print(f"✅ Done! Output written to: {output_path}")
    print(f"   📊 {len(tables)} tables, {sum(len(c) for c in all_columns.values())} columns, {len(slices)} slices, {len(views)} views, {len(actions)} actions, {len(bots)} bots")


if __name__ == '__main__':
    main()
