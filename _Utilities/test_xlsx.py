import openpyxl
wb = openpyxl.load_workbook('/Users/ommnomi/AppSheets/Projects/Transcend/Transcend Main.xlsx')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    headers = [cell.value for cell in ws[1] if cell.value]
    print(f"{sheet}: {headers}")
